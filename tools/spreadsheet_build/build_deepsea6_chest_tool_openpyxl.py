from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


INPUT_PATH = Path("/Users/tadaema/Documents/sqsd/全服宝箱预览6.1.xlsx")
OUTPUT_DIR = Path("/Users/tadaema/Projects/Stone Age_Script/outputs/chest_probability_tool")
OUTPUT_PATH = OUTPUT_DIR / "全服宝箱预览6.1_深海6楼出货统计工具_兼容版.xlsx"

ITEMS = [
    "焰狱魔兽自选",
    "战神斧碎片",
    "暴躁蛮牛自选",
    "技能碎片",
    "魔神石",
    "魔法技能碎片",
    "深海鱼鳞",
    "深海贝壳",
    "深海之泪",
    "深海泥土",
    "战神首饰碎片",
    "暴龙玩具3",
    "暴龙玩具4",
    "10点金币",
    "剧毒宝石碎片",
    "沉默勋章3",
    "粉红暴自选",
    "粉红暴骑证",
    "3D人龙自选",
]

SHEET_ITEMS = "深海6楼_物品"
SHEET_INPUT = "深海6楼_录入"
SHEET_STATS = "深海6楼_统计"

INPUT_ROWS = 5000
INPUT_START = 4
INPUT_END = INPUT_START + INPUT_ROWS - 1
STATS_START = 11
STATS_END = STATS_START + len(ITEMS) - 1

PRIMARY = "0F6B66"
HEADER = "D7ECE8"
PANEL = "F4FAF8"
NOTE = "FFF8EA"
TEXT = "183D3A"
GRID = "D9E6E3"


def reset_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def style_range(ws, cell_range, fill=None, font=None, alignment=None, border=True):
    thin = Side(style="thin", color=GRID)
    border_style = Border(left=thin, right=thin, top=thin, bottom=thin) if border else None
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border_style:
                cell.border = border_style


def title(ws, cell_range, text):
    start = cell_range.split(":")[0]
    ws[start] = text
    ws.merge_cells(cell_range)
    style_range(
        ws,
        cell_range,
        fill=PatternFill("solid", fgColor=PRIMARY),
        font=Font(name="Aptos", size=16, bold=True, color="FFFFFF"),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=False,
    )
    ws.row_dimensions[int("".join(filter(str.isdigit, start)))].height = 24


def headers(ws, cell_range):
    style_range(
        ws,
        cell_range,
        fill=PatternFill("solid", fgColor=HEADER),
        font=Font(name="Aptos", size=11, bold=True, color=TEXT),
        alignment=Alignment(horizontal="center", vertical="center"),
    )


def body_style(ws, cell_range):
    style_range(
        ws,
        cell_range,
        font=Font(name="Aptos", size=11, color="111827"),
        alignment=Alignment(vertical="center"),
    )


def build_workbook():
    wb = load_workbook(INPUT_PATH)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    items_ws = reset_sheet(wb, SHEET_ITEMS)
    input_ws = reset_sheet(wb, SHEET_INPUT)
    stats_ws = reset_sheet(wb, SHEET_STATS)

    for ws in (items_ws, input_ws, stats_ws):
        ws.sheet_view.showGridLines = False

    # 物品清单
    title(items_ws, "A1:C1", "深海6楼箱子物品清单")
    items_ws.append([])
    items_ws.append(["序号", "物品", "备注"])
    headers(items_ws, "A3:C3")
    for idx, item in enumerate(ITEMS, start=1):
        items_ws.append([idx, item, ""])
    body_style(items_ws, f"A4:C{3 + len(ITEMS)}")
    items_ws.column_dimensions["A"].width = 8
    items_ws.column_dimensions["B"].width = 22
    items_ws.column_dimensions["C"].width = 24

    # 录入页
    title(input_ws, "A1:D1", "深海6楼箱子掉落录入")
    input_ws["A2"] = "只填日期、物品、数量；数量合计就是箱子数，同一天自动算一个批次。"
    input_ws.merge_cells("A2:D2")
    style_range(
        input_ws,
        "A2:D2",
        fill=PatternFill("solid", fgColor="EEF7F5"),
        font=Font(name="Aptos", size=11, color="295754"),
        alignment=Alignment(vertical="center"),
        border=False,
    )
    input_ws.append(["日期", "物品", "数量", "备注", "批次标记"])
    headers(input_ws, "A3:E3")
    input_ws.freeze_panes = "A4"
    input_ws.column_dimensions["A"].width = 14
    input_ws.column_dimensions["B"].width = 22
    input_ws.column_dimensions["C"].width = 10
    input_ws.column_dimensions["D"].width = 34
    input_ws.column_dimensions["E"].width = 10
    input_ws.column_dimensions["E"].hidden = True

    item_validation = DataValidation(
        type="list",
        formula1=f"'{SHEET_ITEMS}'!$B$4:$B${3 + len(ITEMS)}",
        allow_blank=True,
    )
    qty_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="999999",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="数量无效",
        error="数量请输入 0 或正整数。",
    )
    input_ws.add_data_validation(item_validation)
    input_ws.add_data_validation(qty_validation)
    item_validation.add(f"B{INPUT_START}:B{INPUT_END}")
    qty_validation.add(f"C{INPUT_START}:C{INPUT_END}")

    for row in range(INPUT_START, INPUT_END + 1):
        input_ws[f"E{row}"] = f'=IF(A{row}="","",IF(COUNTIF($A$4:A{row},A{row})=1,1,0))'
        input_ws[f"A{row}"].number_format = "yyyy-mm-dd"
        input_ws[f"C{row}"].number_format = "0"
    body_style(input_ws, f"A{INPUT_START}:D{INPUT_START + 120}")

    # 统计页
    title(stats_ws, "A1:F1", "深海6楼箱子出货统计")
    summary_rows = [
        ("累计箱子数", f"=SUM('{SHEET_INPUT}'!$C${INPUT_START}:$C${INPUT_END})"),
        ("记录批次数", f"=SUM('{SHEET_INPUT}'!$E${INPUT_START}:$E${INPUT_END})"),
        ("平均每批箱子数", '=IF($B$4>0,$B$3/$B$4,"")'),
        ("明细行数", f"=COUNT('{SHEET_INPUT}'!$C${INPUT_START}:$C${INPUT_END})"),
        (
            "未填日期的明细",
            f'=COUNTIFS(\'{SHEET_INPUT}\'!$B${INPUT_START}:$B${INPUT_END},"<>",\'{SHEET_INPUT}\'!$A${INPUT_START}:$A${INPUT_END},"")',
        ),
        (
            "未填数量的明细",
            f'=COUNTIFS(\'{SHEET_INPUT}\'!$B${INPUT_START}:$B${INPUT_END},"<>",\'{SHEET_INPUT}\'!$C${INPUT_START}:$C${INPUT_END},"")',
        ),
    ]
    for idx, (label, formula) in enumerate(summary_rows, start=3):
        stats_ws[f"A{idx}"] = label
        stats_ws[f"B{idx}"] = formula
    style_range(
        stats_ws,
        "A3:B8",
        fill=PatternFill("solid", fgColor=PANEL),
        font=Font(name="Aptos", size=11, bold=True, color=TEXT),
        alignment=Alignment(vertical="center"),
    )
    for row in range(3, 9):
        stats_ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
        stats_ws[f"B{row}"].number_format = "0.00" if row == 5 else "0"

    notes = [
        "说明",
        "每个箱子只按最终掉落物记录一次，所以数量合计就是开箱数。",
        "同一天多行记录会自动算作 1 个批次。",
        "出货率 = 该物品累计数量 / 累计箱子数。",
    ]
    for offset, value in enumerate(notes, start=3):
        stats_ws[f"D{offset}"] = value
        stats_ws.merge_cells(start_row=offset, start_column=4, end_row=offset, end_column=6)
    style_range(
        stats_ws,
        "D3:F6",
        fill=PatternFill("solid", fgColor=NOTE),
        font=Font(name="Aptos", size=11, color="111827"),
        alignment=Alignment(vertical="center", wrap_text=True),
        border=False,
    )
    stats_ws["D3"].font = Font(name="Aptos", size=11, bold=True, color="111827")

    for col, label in enumerate(["物品", "累计数量", "出货率", "数量排名", "相对条", "备注"], start=1):
        stats_ws.cell(row=10, column=col, value=label)
    headers(stats_ws, "A10:F10")

    for offset, item in enumerate(ITEMS):
        row = STATS_START + offset
        stats_ws[f"A{row}"] = item
        stats_ws[f"B{row}"] = (
            f"=SUMIF('{SHEET_INPUT}'!$B${INPUT_START}:$B${INPUT_END},"
            f"A{row},'{SHEET_INPUT}'!$C${INPUT_START}:$C${INPUT_END})"
        )
        stats_ws[f"C{row}"] = f'=IF($B$3>0,B{row}/$B$3,"")'
        stats_ws[f"D{row}"] = f'=IF(B{row}>0,RANK(B{row},$B${STATS_START}:$B${STATS_END}),"")'
        stats_ws[f"E{row}"] = f'=IFERROR(REPT("|",ROUND(C{row}/MAX($C${STATS_START}:$C${STATS_END})*18,0)),"")'
        stats_ws[f"B{row}"].number_format = "0"
        stats_ws[f"C{row}"].number_format = "0.00%"
        stats_ws[f"D{row}"].number_format = "0"
        stats_ws[f"E{row}"].font = Font(name="Aptos", size=11, color=PRIMARY)
    body_style(stats_ws, f"A{STATS_START}:F{STATS_END}")
    stats_ws.freeze_panes = "A11"
    stats_ws.column_dimensions["A"].width = 22
    stats_ws.column_dimensions["B"].width = 12
    stats_ws.column_dimensions["C"].width = 12
    stats_ws.column_dimensions["D"].width = 10
    stats_ws.column_dimensions["E"].width = 20
    stats_ws.column_dimensions["F"].width = 20

    return wb


if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    # Reopen once to catch malformed workbook structures before handing it over.
    load_workbook(OUTPUT_PATH, data_only=False)
    print(OUTPUT_PATH)
