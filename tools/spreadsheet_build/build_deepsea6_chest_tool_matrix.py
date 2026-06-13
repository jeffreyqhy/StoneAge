from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


INPUT_PATH = Path("/Users/tadaema/Documents/sqsd/全服宝箱预览6.1.xlsx")
OUTPUT_DIR = Path("/Users/tadaema/Projects/Stone Age_Script/outputs/chest_probability_tool")
OUTPUT_PATH = OUTPUT_DIR / "全服宝箱预览6.1_深海6楼出货统计工具_横向加总版.xlsx"

ITEMS = [
    "焰狱魔兽自选",
    "战神斧碎片",
    "暴躁蛮牛自选",
    "技能碎片",
    "魔神石",
    "魔法技能碎片",
    "深海鱼鳞",
    "深海贝壳",
    "深海之泪",
    "深海泥土",
    "战神首饰碎片",
    "暴龙玩具3",
    "暴龙玩具4",
    "10点金币",
    "剧毒宝石碎片",
    "沉默勋章3",
    "粉红暴自选",
    "粉红暴骑证",
    "3D人龙自选",
]

SHEET_ITEMS = "深海6楼_物品"
SHEET_INPUT = "深海6楼_快速录入"
SHEET_STATS = "深海6楼_统计"

RECORD_COLS = 200
FIRST_RECORD_COL = 3
LAST_RECORD_COL = FIRST_RECORD_COL + RECORD_COLS - 1
LAST_RECORD_LETTER = get_column_letter(LAST_RECORD_COL)
ITEM_START_ROW = 14
ITEM_END_ROW = ITEM_START_ROW + len(ITEMS) - 1

PRIMARY = "0F6B66"
HEADER = "D7ECE8"
PANEL = "F4FAF8"
NOTE = "FFF8EA"
TEXT = "183D3A"
GRID = "D9E6E3"
WARN = "FFF1D6"


def reset_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def style_range(ws, cell_range, fill=None, font=None, alignment=None, border=True):
    thin = Side(style="thin", color=GRID)
    border_style = Border(left=thin, right=thin, top=thin, bottom=thin) if border else None
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border_style:
                cell.border = border_style


def title(ws, cell_range, text):
    start = cell_range.split(":")[0]
    ws[start] = text
    ws.merge_cells(cell_range)
    style_range(
        ws,
        cell_range,
        fill=PatternFill("solid", fgColor=PRIMARY),
        font=Font(name="Aptos", size=16, bold=True, color="FFFFFF"),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=False,
    )
    ws.row_dimensions[int("".join(filter(str.isdigit, start)))].height = 24


def headers(ws, cell_range):
    style_range(
        ws,
        cell_range,
        fill=PatternFill("solid", fgColor=HEADER),
        font=Font(name="Aptos", size=11, bold=True, color=TEXT),
        alignment=Alignment(horizontal="center", vertical="center"),
    )


def body_style(ws, cell_range):
    style_range(
        ws,
        cell_range,
        font=Font(name="Aptos", size=11, color="111827"),
        alignment=Alignment(vertical="center"),
    )


def set_calc_mode(wb):
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass


def build_items_sheet(ws):
    title(ws, "A1:C1", "深海6楼箱子物品清单")
    ws.append([])
    ws.append(["序号", "物品", "备注"])
    headers(ws, "A3:C3")
    for idx, item in enumerate(ITEMS, start=1):
        ws.append([idx, item, ""])
    body_style(ws, f"A4:C{3 + len(ITEMS)}")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24


def build_input_sheet(ws):
    title(ws, "A1:O1", "深海6楼箱子快速录入")
    ws["A2"] = "左边固定物品；右边每一列代表一个号/一次5箱。开完一组就去下一列填数字，不用回头改旧数量。"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    style_range(
        ws,
        "A2:O2",
        fill=PatternFill("solid", fgColor="EEF7F5"),
        font=Font(name="Aptos", size=11, color="295754"),
        alignment=Alignment(vertical="center"),
        border=False,
    )

    ws["A3"] = "汇总"
    ws["B3"] = "数值"
    ws["C3"] = "每列 = 一个号 / 一次5箱"
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=15)
    ws["A4"] = "累计箱子数"
    ws["B4"] = f"=SUM(B{ITEM_START_ROW}:B{ITEM_END_ROW})"
    ws["A5"] = "开箱列数"
    ws["B5"] = f'=COUNTIF(C11:{LAST_RECORD_LETTER}11,">0")'
    ws["A6"] = "平均每列箱数"
    ws["B6"] = '=IF(B5>0,B4/B5,"")'
    ws["A7"] = "非5箱列数"
    ws["B7"] = f'=COUNTIF(C12:{LAST_RECORD_LETTER}12,"检查")'
    headers(ws, "A3:B3")
    headers(ws, "C3:O3")
    style_range(
        ws,
        "A4:B7",
        fill=PatternFill("solid", fgColor=PANEL),
        font=Font(name="Aptos", size=11, bold=True, color=TEXT),
        alignment=Alignment(vertical="center"),
    )
    for row in range(4, 8):
        ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B6"].number_format = "0.00"

    for row, label in {
        8: "记录号",
        9: "日期",
        10: "账号/备注",
        11: "本列箱数",
        12: "状态",
        13: "物品",
    }.items():
        ws[f"A{row}"] = label
    ws["B13"] = "累计数量"
    for col in range(FIRST_RECORD_COL, LAST_RECORD_COL + 1):
        letter = get_column_letter(col)
        record_num = col - FIRST_RECORD_COL + 1
        ws[f"{letter}8"] = record_num
        ws[f"{letter}9"] = ""
        ws[f"{letter}10"] = ""
        ws[f"{letter}11"] = f"=SUM({letter}{ITEM_START_ROW}:{letter}{ITEM_END_ROW})"
        ws[f"{letter}12"] = f'=IF({letter}11=0,"",IF({letter}11=5,"OK","检查"))'
        ws[f"{letter}13"] = record_num
        ws[f"{letter}9"].number_format = "yyyy-mm-dd"
        ws[f"{letter}11"].number_format = "0"
        ws[f"{letter}12"].alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[letter].width = 7

    headers(ws, f"A8:B13")
    headers(ws, f"A13:{LAST_RECORD_LETTER}13")
    style_range(
        ws,
        f"C8:{LAST_RECORD_LETTER}12",
        fill=PatternFill("solid", fgColor="F8FBFA"),
        font=Font(name="Aptos", size=10, color="111827"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    headers(ws, f"C13:{LAST_RECORD_LETTER}13")

    for offset, item in enumerate(ITEMS):
        row = ITEM_START_ROW + offset
        ws[f"A{row}"] = item
        ws[f"B{row}"] = f"=SUM(C{row}:{LAST_RECORD_LETTER}{row})"
        ws[f"B{row}"].number_format = "0"

    body_style(ws, f"A{ITEM_START_ROW}:{LAST_RECORD_LETTER}{ITEM_END_ROW}")
    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        ws[f"A{row}"].font = Font(name="Aptos", size=11, bold=True, color="111827")
        ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")

    qty_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="999999",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="数量无效",
        error="数量请输入 0 或正整数。",
    )
    ws.add_data_validation(qty_validation)
    qty_validation.add(f"C{ITEM_START_ROW}:{LAST_RECORD_LETTER}{ITEM_END_ROW}")

    ws.freeze_panes = f"C{ITEM_START_ROW}"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    for row in range(1, ITEM_END_ROW + 1):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22


def build_stats_sheet(ws):
    title(ws, "A1:F1", "深海6楼箱子出货统计")
    summary = [
        ("累计箱子数", f"='{SHEET_INPUT}'!$B$4"),
        ("开箱列数", f"='{SHEET_INPUT}'!$B$5"),
        ("平均每列箱数", f"='{SHEET_INPUT}'!$B$6"),
        ("非5箱列数", f"='{SHEET_INPUT}'!$B$7"),
    ]
    for idx, (label, formula) in enumerate(summary, start=3):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = formula
    style_range(
        ws,
        "A3:B6",
        fill=PatternFill("solid", fgColor=PANEL),
        font=Font(name="Aptos", size=11, bold=True, color=TEXT),
        alignment=Alignment(vertical="center"),
    )
    for row in range(3, 7):
        ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B5"].number_format = "0.00"

    notes = [
        "说明",
        "快速录入页每一列代表一个号/一次5箱。",
        "同一物品在不同列填多次，会自动横向加总。",
        "非5箱列数不为0时，说明有某列合计不是5，建议回录入页检查。",
    ]
    for offset, value in enumerate(notes, start=3):
        ws[f"D{offset}"] = value
        ws.merge_cells(start_row=offset, start_column=4, end_row=offset, end_column=6)
    style_range(
        ws,
        "D3:F6",
        fill=PatternFill("solid", fgColor=NOTE),
        font=Font(name="Aptos", size=11, color="111827"),
        alignment=Alignment(vertical="center", wrap_text=True),
        border=False,
    )
    ws["D3"].font = Font(name="Aptos", size=11, bold=True, color="111827")

    for col, label in enumerate(["物品", "累计数量", "出货率", "数量排名", "相对条", "备注"], start=1):
        ws.cell(row=9, column=col, value=label)
    headers(ws, "A9:F9")

    for offset, item in enumerate(ITEMS):
        row = 10 + offset
        input_row = ITEM_START_ROW + offset
        ws[f"A{row}"] = item
        ws[f"B{row}"] = f"='{SHEET_INPUT}'!$B${input_row}"
        ws[f"C{row}"] = f'=IF($B$3>0,B{row}/$B$3,"")'
        ws[f"D{row}"] = f'=IF(B{row}>0,RANK(B{row},$B$10:$B${9 + len(ITEMS)}),"")'
        ws[f"E{row}"] = f'=IFERROR(REPT("|",ROUND(C{row}/MAX($C$10:$C${9 + len(ITEMS)})*18,0)),"")'
        ws[f"B{row}"].number_format = "0"
        ws[f"C{row}"].number_format = "0.00%"
        ws[f"D{row}"].number_format = "0"
        ws[f"E{row}"].font = Font(name="Aptos", size=11, color=PRIMARY)
    body_style(ws, f"A10:F{9 + len(ITEMS)}")
    ws.freeze_panes = "A10"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 24


def build_workbook():
    wb = load_workbook(INPUT_PATH)
    set_calc_mode(wb)

    items_ws = reset_sheet(wb, SHEET_ITEMS)
    input_ws = reset_sheet(wb, SHEET_INPUT)
    stats_ws = reset_sheet(wb, SHEET_STATS)

    for ws in (items_ws, input_ws, stats_ws):
        ws.sheet_view.showGridLines = False

    build_items_sheet(items_ws)
    build_input_sheet(input_ws)
    build_stats_sheet(stats_ws)
    return wb


if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    load_workbook(OUTPUT_PATH, data_only=False)
    print(OUTPUT_PATH)
